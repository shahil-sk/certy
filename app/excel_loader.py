"""
Data ingestion -- supports .xlsx and .csv.
Returns plain Python dicts; caller owns the UI feedback.
"""
import csv
from datetime import datetime

from openpyxl import load_workbook

from app.logger import get_logger

log = get_logger(__name__)


def read(file_path: str) -> tuple[list, list]:
    """
    Read an xlsx or csv file and return (header, rows).
    header : list of lowercase, stripped column names
    rows   : list of dicts  {col_name: str_value}
    Raises ValueError on bad input.
    """
    log.debug("reading data file: %s", file_path)
    if file_path.lower().endswith(".csv"):
        return _read_csv(file_path)
    return _read_xlsx(file_path)


def sheet_names(file_path: str) -> list[str]:
    """
    Return the list of sheet names in an xlsx workbook.
    Returns an empty list for CSV files (they have no sheets).
    """
    if file_path.lower().endswith(".csv"):
        return []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_sheet(file_path: str, sheet: str) -> tuple[list, list]:
    """
    Read a specific named sheet from an xlsx workbook.
    Falls back to the active sheet if sheet name is not found.
    """
    wb    = load_workbook(file_path, read_only=True, data_only=True)
    names = wb.sheetnames
    if sheet in names:
        ws = wb[sheet]
    else:
        log.warning("sheet '%s' not found, falling back to active sheet", sheet)
        ws = wb.active

    header, rows = _parse_sheet(ws, file_path)
    wb.close()
    return header, rows


def _read_xlsx(file_path: str) -> tuple[list, list]:
    wb    = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active
    header, rows = _parse_sheet(sheet, file_path)
    wb.close()
    return header, rows


def _parse_sheet(ws, file_path: str) -> tuple[list, list]:
    """Extract (header, rows) from an already-opened worksheet object."""
    header = [
        str(c.value).strip().lower()
        for c in next(ws.iter_rows(min_row=1, max_row=1))
        if c.value is not None
    ]
    if not header:
        raise ValueError("No column headers found in the selected sheet.")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for i, val in enumerate(row):
            if i >= len(header):
                break
            if isinstance(val, datetime):
                val = val.strftime("%d-%m-%Y")
            rec[header[i]] = str(val) if val is not None else ""
        if any(rec.values()):
            rows.append(rec)

    if not rows:
        raise ValueError("No data rows found in the selected sheet.")
    log.info("sheet loaded: %d rows, %d columns from '%s'", len(rows), len(header), file_path)
    return header, rows


def _read_csv(file_path: str) -> tuple[list, list]:
    rows   = []
    header = []
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(file_path, newline="", encoding=enc) as fh:
                reader = csv.DictReader(fh)
                if not reader.fieldnames:
                    raise ValueError("No column headers found in the CSV file.")
                header = [f.strip().lower() for f in reader.fieldnames]
                for row in reader:
                    rec = {h: str(row.get(orig, "") or "").strip()
                           for h, orig in zip(header, reader.fieldnames)}
                    if any(rec.values()):
                        rows.append(rec)
            log.info("csv loaded (enc=%s): %d rows, %d columns from '%s'",
                     enc, len(rows), len(header), file_path)
            break
        except UnicodeDecodeError:
            log.debug("encoding %s failed for '%s', trying next", enc, file_path)
            continue

    if not header:
        raise ValueError("Could not decode CSV file.")
    if not rows:
        raise ValueError("No data rows found in the CSV file.")
    return header, rows
